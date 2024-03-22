import openpyxl

fpath= r'C:\Users\kio23\Desktop\front\py_rolli\3.파이썬엑셀다루기\참가자_데이터.xlsx'

# 1. 엑셀 파일 불러오기
wb = openpyxl.load_workbook(fpath)

# 2. 엑셀 시트 선택
ws = wb['오징어게임']

#3. 데이터 수정
ws['A3'] = 456
ws['B3'] = '성기훈'

#4. 엑셀 저장
wb.save(fpath)