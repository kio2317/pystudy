import openpyxl

# 1.엑셀 만들기
wb = openpyxl.Workbook()

# 2. 워크 시트 만들기
ws = wb.create_sheet('오징어게임')

#3. 데이터 추가
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'
# 4. 엑셀 저장
wb.save(r'C:\Users\kio23\Desktop\front\py_rolli\3.파이썬엑셀다루기\참가자_데이터.xlsx')
